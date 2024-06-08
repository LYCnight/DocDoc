from pathlib import Path		
import sys
cur_path = Path(__file__).parent    # 当前目录    DocDoc/test/GPT4_test
test_path = Path(__file__).parent.parent    # 测试目录    DocDoc/test
root_path = Path(__file__).parent.parent.parent    # 项目根目录    DocDoc/
sys.path.append(str(cur_path))
sys.path.append(str(test_path))
sys.path.append(str(root_path))

import openpyxl
from openpyxl import load_workbook

from DocDoc_EMNLP import DocDoc_write

def write(prompt:str, title:str) -> str:
    """ generate a text based on the prompt"""
    print(f"I am writing <{title}>")
    # write the text to a file
    file_path:str = DocDoc_write(prompt)
    # -- for test -- 
    # file_path = str(cur_path) + "/output/DocDoc/" + title + ".md"   
    # with open(file_path, "w") as f:
    #     f.write(prompt)
    # --------------
    print(f"{title} has been written to {file_path}")
    return file_path


class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename
        try:
            # Try to load an existing workbook
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
        except FileNotFoundError:
            # If the file doesn't exist, create a new one
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(["id", "Title", "Title_ch", "Category", "Type", "Strategy", "Prompt", "Plan", "Text"])  # Add headers if it's a new file

    def __iter__(self):
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            yield dict(zip([cell.value for cell in self.ws[1]], row))
    
    def update_row(self, row_num, col_name, value):
        col_idx = [cell.value for cell in self.ws[1]].index(col_name) + 1
        self.ws.cell(row=row_num, column=col_idx, value=value)

    def save(self):
        self.wb.save(self.filename)

def make_prompt(title:str, category:str) -> str:
    prompt:str = f"""I want to write one {category}, titled "{title}" Could you generate the table of contents for the opinion article and provide a detailed explanation of the dependencies between the items in the table of contents?"""
    return prompt


if __name__ == "__main__":
# - test for make_prompt" 
    # print(make_prompt("AAa", "BBB"))

    xlsx_file_path:str =  "/root/AI4E/lzd/DocDoc/test/GPT4_test/output_refined_1.xlsx"
    excel = ExcelHandler(xlsx_file_path)

    # Iterate through each row in the Excel file
    for row_num, row in enumerate(excel, start=2):
        if(row['Done'] != 1):   # 还没有被处理过
            title: str = row['Title']
            category: str = row['Category']
            prompt: str = make_prompt(title, category)
            file_path: str = write(prompt, title)
            excel.update_row(row_num, 'Prompt', prompt)
            excel.update_row(row_num, 'Text', file_path)
            excel.update_row(row_num, 'Done', 1)    # if done, set to 1
            # Save the updated Excel file
            excel.save()        # < 我可以这样写 从而实现边写边存吗？这样就算程序崩溃我也可以获得一部分结果>

    # Save the updated Excel file
    excel.save()

    print("completed successfully!")

    # -- test --
    # """write one doc, count token used, money expenditure and time"""
    # title = """The Museum Murder Mystery"""
    # prompt = """I want to write one Mystery Fiction, titled "The Museum Murder Mystery" Could you generate the table of contents for the opinion article and provide a detailed explanation of the dependencies between the items in the table of contents?"""
    # file_path = write(prompt=prompt, title=title)
    # print(f"successed! file_path: {file_path}")

        