prompt_template = """<rule>
All texts can be classified based on their directory depth:
- Shallow: The directory structure is relatively shallow, with levels ranging from 0 to 1. It involves linear narratives with only one level of directory items, excluding multi-level directory items.
    - Fiction
        - Adventure Fiction
        - Romance Fiction
        - Science Fiction
        - Fantasy Fiction
        - Horror Fiction
        - Thriller Fiction
        - Mystery Fiction
        - Historical Fiction
        - Drama
    - News
        - Political News
        - Business and Economy News
        - Technology News
        - Sports News
        - Entertainment News
        - Science and Health News
        - Environment News
    - Opinion articles or review
        - Political Opinion
        - Economic Opinion
        - Social Commentary
        - Satire and Humor Opinion
        - Movie review or Book review
- Medium: The directory structure has a moderate depth, with levels ranging from 0 to 3, including multi-level directory items.
    - Academic papers
        - Physics Paper
        - Chemistry Paper
        - Biology Paper
        - Environment Paper
        - Computer Science Paper
        - Artificial Intelligence Paper
        - Philosophy Paper
        - Psychology Paper
        - History Paper
        - Sociology Paper
    - Encyclopedia articles
        - a historical figure's encyclopedia articles
        - a historical event's encyclopedia articles
        - a terminology's encyclopedia articles
        - a computer game's encyclopedia articles
        - a company's encyclopedia articles
        - a country's encyclopedia articles
- Deep: The directory structure is very deep, with levels ranging from 0 to 6, including deeply nested directory items.
    - Report
        - IT Report: Software Development Report and more
        - Medicine Report: Clinical Study Report and more
        - Finance Report: Risk Assessment Report and more
        - Education Report: Course Evaluation Report and more
        - Law Report: Case Assessment Report and more
        - Management Report: Project Management Report and more
        - Manufacturing Report: Manufacturing Process Report and more
        - Environment Report: Environmental Assessment Report and more
    - TextBook
        - Computer Science Textbook: `Data Sturctures` or `Python Tutorial` and more
        - English Textbook: `Grammar` or `English Writing` and more
        - History Textbook: `World History` or `American History` and more
</rule>
<attention>
1. the titles you generated must be with id.
2. you must wrap your titles with <title> and </title>
</attention>
Q: Could you please provide me with 10 titles for Fiction classified as Shallow?
A:
<title>
1.  "The Mystery of Blackwood Manor"
2. "A Journey Through Time"
3. "The Secret Garden"
4. "Lost in the Woods"
5. "The Haunting of Hill House"
6. "The Adventures of Captain Redbeard"
7. "A Love Story in Paris"
8. "The Ghostly Visitor"
9. "The Enchanted Forest"
10. "Whispers in the Wind"
</title>

Q: Could you please provide me with {generate_per_round} titles for {cate} classified as {type}?
A: 
"""



from pathlib import Path		
import sys
cur_path = Path(__file__).parent    # 当前目录    DocDoc/test/GPT4_test
test_path = Path(__file__).parent.parent    # 测试目录    DocDoc/test
root_path = Path(__file__).parent.parent.parent    # 项目根目录    DocDoc/
sys.path.append(str(cur_path))
sys.path.append(str(test_path))
sys.path.append(str(root_path))

# from prompt import prompt1
from core.Agent import ContentExpert
from DocDoc import Heading
from LLMs import ChatGLM

import re
import openpyxl
from openpyxl import load_workbook
class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename
        try:
            # Try to load an existing workbook
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
        except FileNotFoundError:
            # If the file doesn't exist, create a new one
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(["Title", "Category", "Type"])  # Add headers if it's a new file

    def persist(self, title, cate, k):
        self.ws.append([title, cate, k])

    def save(self):
        self.wb.save(self.filename)

def extract_title(text: str) -> list[str]:
    # Find the section within the <title> tags
    start_tag = "<title>"
    end_tag = "</title>"
    start_index = text.find(start_tag) + len(start_tag)
    end_index = text.find(end_tag)
    title_section = text[start_index:end_index].strip()
    # Split the title section into lines
    lines = title_section.split('\n')
    # Extract the title from each line
    titles = [line.split('.', 1)[1].strip().strip('"') for line in lines if '.' in line]
    return titles

def read_excel(file_path):
    '''
    读取xlsx，转换成这样的dict:
    type_dict = {"Shallow": [{"Math": 1}, {"Physics": 2}],
                "Medium": [{"Chemistry": 3}, {"Biology": 4}],
                "Deep": [{"History": 5}, {"Geography": 6}]}
    '''
    import pandas as pd
    # 读取 Excel 文件
    df = pd.read_excel(file_path)

    # 构建目标字典
    type_dict = {}
    for _, row in df.iterrows():
        type_ = row['type'].capitalize()
        category = row['category']
        count = row['count']
        
        if type_ not in type_dict:
            type_dict[type_] = []
        
        type_dict[type_].append({category: count})
        
    # test: 打印字典
    # print(type_dict)

    # test: 格式化打印目标字典
    # for type_, categories in type_dict.items():
    #     print(f"{type_}:")
    #     for category in categories:
    #         print(f"  {list(category.keys())[0]}: {list(category.values())[0]}")
    #     print()

    return type_dict

# -------------------------------------- main ---------------------------------------
file_path = str(cur_path) +"/category.xlsx"
print(file_path)
type_dict:dict = read_excel(file_path)

# -- test fpr type_dict --
# type_dict = {"Shallow": [{"Math": 1}, {"Physics": 2}],
#              "Medium": [{"Chemistry": 3}, {"Biology": 4}],
#              "Deep": [{"History": 5}, {"Geography": 6}]}

# -- test for extract_title --
# text = """
# 1.  "The Mystery of Blackwood Manor"
# 2. "A Journey Through Time"
# 3. "The Secret Garden"
# 4. "Lost in the Woods"
# 5. "The Haunting of Hill House"
# 6. "The Adventures of Captain Redbeard"
# 7. "A Love Story in Paris"
# 8. "The Ghostly Visitor"
# 9. "The Enchanted Forest"
# 10. "Whispers in the Wind"
# """
# title_list = extract_title(text)
# print(title_list)

# Initialize llm and Excel handler
llm = ChatGLM()
output_path = str(cur_path) +"/output.xlsx"
excel = ExcelHandler(output_path)
# print(file_path)

# -- test for llm --
# def llm(k, cate=None):
#     if cate:
#         return f'Titles for {cate} in {k} category:\n1. "Example Title 1"\n2. "Example Title 2"'
#     else:
#         return 'Example response from LLM'

for k, v in type_dict.items():
    for item in v:
        for cate, count in item.items():  # Use .items() to unpack key-value pairs
            title_list:list[str] = []
            while(len(title_list) < count):
                prompt = prompt_template.format(generate_per_round=count, cate=cate, type=k)
                print(prompt)
                print("*"*100)
                response:str = llm(prompt)
                print(response)
                # response = llm(k,cate)
                title_list.extend(extract_title(response))
            title_list = title_list[:count]
            # persis to excel
            for title in title_list:
                excel.persist(title, cate, k)

excel.save()


print(file_path)
print(str(cur_path) +"output.xlsx")


# 单次生成：
# from LLMs import ChatGLM
# llm = ChatGLM()
# prompt = prompt_template.format(generate_per_round=2, cate="Artificial Intelligence Paper", type="Medium")
# response = llm(prompt)
# print(response)