import re
import openpyxl
from openpyxl import load_workbook

def extract_title(text: str) -> list[str]:
    # Split the text into lines
    lines = text.strip().split('\n')
    
    titles = []
    for line in lines:
        # Use regular expression to find titles within quotes
        match = re.search(r'\"(.*?)\"', line)
        if match:
            titles.append(match.group(1).strip())
    
    return titles

class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename
        try:
            # Try to load an existing workbook
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
        except FileNotFoundError:
            # If the file doesn't exist, create a new one
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(["Title", "Category", "Type"])  # Add headers if it's a new file

    def persist(self, title, cate, k):
        self.ws.append([title, cate, k])

    def save(self):
        self.wb.save(self.filename)

# Placeholder function for the language model
# def llm(k, cate=None):
#     if cate:
#         return f'Titles for {cate} in {k} category:\n1. "Example Title 1"\n2. "Example Title 2"'
#     else:
#         return 'Example response from LLM'

type_dict = {"Shallow": [{"Math": 1}, {"Physics": 2}],
             "Medium": [{"Chemistry": 3}, {"Biology": 4}],
             "Deep": [{"History": 5}, {"Geography": 6}]}

# Initialize Excel handler
excel = ExcelHandler("output.xlsx")

for k, v in type_dict.items():
    for item in v:
        for cate, count in item.items():  # Use .items() to unpack key-value pairs
            title_list = []
            while len(title_list) < count:
                response = llm(k, cate)
                title_list.extend(extract_title(response))
            title_list = title_list[:count]
            # Persist to excel
            for title in title_list:
                excel.persist(title, cate, k)   # 保存一行数据

# Save the excel file
excel.save()
